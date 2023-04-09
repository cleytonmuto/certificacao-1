import pandas
import openpyxl

class SheetLoader:

    def __init__(self):
        pass

    def loadSistemas(self, path):
        df = pandas.read_excel(path, sheet_name="sistemas")
        codigo, sistema = [], []
        for elem in df["codigo"]:
            codigo.append(elem)
        for elem in df["sistema"]:
            sistema.append(elem)
        return codigo, sistema

    def loadPerfis(self, path):
        df = pandas.read_excel(path, sheet_name="perfis")
        codigo, perfil, descricao = [], [], []
        for elem in df["codigo"]:
            codigo.append(elem)
        for elem in df["perfil"]:
            perfil.append(elem)
        for elem in df["descricao"]:
            descricao.append(elem)
        return codigo, perfil, descricao

    def addSistema(self, path, codigo, sistema):
        df = pandas.read_excel(path, sheet_name="sistemas")
        codigos, sistemas = [], []
        for elem in df["codigo"]:
            codigos.append(elem)
        for elem in df["sistema"]:
            sistemas.append(elem)
        rows = []
        for i in range(len(codigos)):
            rows.append([codigos[i], sistemas[i]])
        rows.append([codigo, sistema])
        output = pandas.DataFrame(rows, columns=["codigo","sistema"])

        book = openpyxl.load_workbook(path)
        writer = pandas.ExcelWriter(path, engine="openpyxl", mode="a",
            if_sheet_exists="overlay")
        writer.workbook = book
        output.to_excel(writer, index=False, sheet_name="sistemas")
        writer.close()
    
    def addPerfil(self, path, codigo, perfil, descricao):
        df = pandas.read_excel(path, sheet_name="perfis")
        codigos, perfis, descricoes = [], [], []
        for elem in df["codigo"]:
            codigos.append(elem)
        for elem in df["perfil"]:
            perfis.append(elem)
        for elem in df["descricao"]:
            descricoes.append(elem)
        rows = []
        for i in range(len(codigos)):
            rows.append([codigos[i], perfis[i], descricoes[i]])
        rows.append([codigo, perfil, descricao])
        output = pandas.DataFrame(rows, columns=["codigo","perfil","descricao"])

        book = openpyxl.load_workbook(path)
        writer = pandas.ExcelWriter(path, engine="openpyxl", mode="a",
            if_sheet_exists="overlay")
        writer.workbook = book
        output.to_excel(writer, index=False, sheet_name="perfis")
        writer.close()
    
    def delSistema(self, path, codigo, sistema):
        df = pandas.read_excel(path, sheet_name="sistemas")
        codigos, sistemas = [], []
        for elem in df["codigo"]:
            codigos.append(elem)
        for elem in df["sistema"]:
            sistemas.append(elem)
        rows = []
        for i in range(len(codigos)):
            rows.append([codigos[i], sistemas[i]])
        for i in range(len(codigos)):
            if codigos[i] == codigo and sistemas[i] == sistema:
                codigos.pop(i)
                sistemas.pop(i)
                rows.pop(i)
                break
        output = pandas.DataFrame(rows, columns=["codigo","sistema"])
        book = openpyxl.load_workbook(path)
        writer = pandas.ExcelWriter(path, engine="openpyxl", mode="a",
            if_sheet_exists="replace")
        writer.workbook = book
        output.to_excel(writer, index=False, sheet_name="sistemas")
        writer.close()
    
    def delPerfil(self, path, codigo, perfil):
        df = pandas.read_excel(path, sheet_name="perfis")
        codigos, perfis, descricoes = [], [], []
        for elem in df["codigo"]:
            codigos.append(elem)
        for elem in df["perfil"]:
            perfis.append(elem)
        for elem in df["descricao"]:
            descricoes.append(elem)
        rows = []
        for i in range(len(codigos)):
            rows.append([codigos[i], perfis[i], descricoes[i]])
        for i in range(len(codigos)):
            if codigos[i] == codigo and perfis[i] == perfil:
                codigos.pop(i)
                perfis.pop(i)
                descricoes.pop(i)
                rows.pop(i)
                break
        output = pandas.DataFrame(rows, columns=["codigo","perfil","descricao"])
        book = openpyxl.load_workbook(path)
        writer = pandas.ExcelWriter(path, engine="openpyxl", mode="a",
            if_sheet_exists="replace")
        writer.workbook = book
        output.to_excel(writer, index=False, sheet_name="perfis")
        writer.close()

if __name__ == "__main__":
    loader = SheetLoader()
    # loader.addSistema("database.xlsx", 4, "madrugada")
    # loader.delSistema("database.xlsx", 2, "tarde")
    loader.addPerfil("database.xlsx", 3, "fiscal", "Fiscal F")
    print("done.")
